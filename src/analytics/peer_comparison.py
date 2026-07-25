import sqlite3
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------------------------------
# Configuration
# -----------------------------------------------------

DATABASE = "db/nifty100.db"
OUTPUT_FILE = "output/peer_comparison.xlsx"

# Header Style
header_fill = PatternFill(fill_type="solid", start_color="1F4E78")
header_font = Font(color="FFFFFF", bold=True)

# Top Rank Style
gold_fill = PatternFill(fill_type="solid", start_color="FFD966")
silver_fill = PatternFill(fill_type="solid", start_color="D9D9D9")
bronze_fill = PatternFill(fill_type="solid", start_color="F4B183")
# -----------------------------------------------------
# Connect to Database
# -----------------------------------------------------

print("Connecting to database...")

conn = sqlite3.connect(DATABASE)

query = """
SELECT
    p.company_id,
    c.company_name,
    p.peer_group_name,
    p.metric,
    p.value,
    p.percentile_rank,
    p.year
FROM peer_percentiles p
JOIN companies c
ON p.company_id = c.id
ORDER BY
    p.peer_group_name,
    c.company_name,
    p.metric;
"""

df = pd.read_sql_query(query, conn)

print("\nData Loaded Successfully")
print(f"Total Records : {len(df)}")
print(f"Peer Groups   : {df['peer_group_name'].nunique()}")
print(f"Companies     : {df['company_id'].nunique()}")

print("\nFirst 5 Rows")
print(df.head())
# -----------------------------------------------------
# Create Workbook
# -----------------------------------------------------

print("\nCreating Excel Workbook...")

wb = Workbook()

# Remove the default sheet
wb.remove(wb.active)
# -----------------------------------------------------
# Generate One Worksheet Per Peer Group
# -----------------------------------------------------

peer_groups = sorted(df["peer_group_name"].unique())

for group in peer_groups:

    print(f"Processing : {group}")

    group_df = df[df["peer_group_name"] == group]

    pivot_df = group_df.pivot_table(
        index="company_name",
        columns="metric",
        values="value",
        aggfunc="first"
    )

    ws = wb.create_sheet(title=group[:31])

    # Header
    headers = ["Company"] + list(pivot_df.columns)

    ws.append(headers)

    # Style Header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data Rows
    for company, row in pivot_df.iterrows():
        ws.append([company] + row.fillna("").tolist())

    # Auto Width
    for column in ws.columns:

        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 3

    ws.freeze_panes = "A2"
    # -----------------------------------------------------
# Save Workbook
# -----------------------------------------------------

wb.save(OUTPUT_FILE)

conn.close()

print("\n--------------------------------")
print("Peer Comparison Report Created")
print("--------------------------------")
print(f"Saved to : {OUTPUT_FILE}")