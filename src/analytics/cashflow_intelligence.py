import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


DATABASE = "db/nifty100.db"
OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "cashflow_intelligence.xlsx"


def load_cashflow_data():
    conn = sqlite3.connect(DATABASE)

    query = """
    SELECT
        c.company_name,
        fr.company_id,
        fr.year,
        fr.cash_from_operations_cr,
        fr.free_cash_flow_cr,
        fr.capex_cr,
        fr.total_debt_cr
    FROM financial_ratios fr
    JOIN companies c
        ON fr.company_id = c.id
    WHERE fr.id IN (
        SELECT MAX(id)
        FROM financial_ratios
        GROUP BY company_id
    )
    ORDER BY c.company_name;
    """

    df = pd.read_sql_query(query, conn)
    conn.close()

    return df

def classify_cashflow(row):
    """
    Classify cash flow health.
    """

    cfo = row["cash_from_operations_cr"] or 0
    fcf = row["free_cash_flow_cr"] or 0
    debt = row["total_debt_cr"] or 0

    if cfo > 0 and fcf > 0 and debt < cfo:
        return "Excellent"

    if cfo > 0 and fcf > 0:
        return "Healthy"

    if cfo > 0:
        return "Average"

    return "Weak"

def export_to_excel(df):
    """Export cashflow intelligence report to Excel."""

    OUTPUT_DIR.mkdir(exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cashflow Intelligence"

    headers = [
        "Company",
        "Year",
        "Cash From Operations (Cr)",
        "Free Cash Flow (Cr)",
        "CapEx",
        "Total Debt (Cr)",
        "Cashflow Health",
    ]

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid",
    )

    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in df.itertuples(index=False):
        ws.append([
            row.company_name.strip(),
            row.year,
            row.cash_from_operations_cr,
            row.free_cash_flow_cr,
            row.capex_cr,
            row.total_debt_cr,
            row.cashflow_health,
        ])

    # Auto-size columns
    for column in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(length + 3, 35)

    ws.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)

    print(f"\nCashflow Intelligence report saved to: {OUTPUT_FILE}")

if __name__ == "__main__":

    df = load_cashflow_data()

    df["cashflow_health"] = df.apply(classify_cashflow, axis=1)

    export_to_excel(df)

    print(
        df[
            [
                "company_name",
                "cashflow_health",
            ]
        ].head()
    )

    print(f"\nCompanies : {len(df)}")