import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DATABASE = "db/nifty100.db"
OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "valuation_summary.xlsx"


def get_latest_financial_ratios():
    """Fetch the latest available financial ratios for each company."""

    conn = sqlite3.connect(DATABASE)

    query = """
    SELECT
        c.company_name,
        fr.company_id,
        fr.year,
        fr.net_profit_margin_pct,
        fr.operating_profit_margin_pct,
        fr.return_on_equity_pct,
        fr.debt_to_equity,
        fr.interest_coverage,
        fr.asset_turnover,
        fr.free_cash_flow_cr,
        fr.earnings_per_share,
        fr.book_value_per_share,
        fr.dividend_payout_ratio_pct
    FROM financial_ratios fr
    JOIN companies c
        ON fr.company_id = c.id
    WHERE fr.id IN (
        SELECT MAX(id)
        FROM financial_ratios
        GROUP BY company_id
    )
    ORDER BY c.company_name;
    """

    df = pd.read_sql_query(query, conn)
    conn.close()

    return df



def calculate_valuation_score(df):
    """Calculate normalized valuation score (0-100)."""

    metrics = [
        "return_on_equity_pct",
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "asset_turnover",
        "interest_coverage",
        "debt_to_equity",
    ]

    data = df.copy()

    for col in metrics:
        data[col] = data[col].fillna(data[col].median())

    positive = [
        "return_on_equity_pct",
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "asset_turnover",
        "interest_coverage",
    ]

    for col in positive:
        mn = data[col].min()
        mx = data[col].max()
        if mx > mn:
            data[col] = (data[col] - mn) / (mx - mn)
        else:
            data[col] = 0.5

    col = "debt_to_equity"
    mn = data[col].min()
    mx = data[col].max()
    if mx > mn:
        data[col] = (mx - data[col]) / (mx - mn)
    else:
        data[col] = 0.5

    score = (
        data["return_on_equity_pct"] * 30
        + data["net_profit_margin_pct"] * 20
        + data["operating_profit_margin_pct"] * 20
        + data["asset_turnover"] * 10
        + data["interest_coverage"] * 10
        + data["debt_to_equity"] * 10
    )

    df["valuation_score"] = score.round(2)

    return df


def assign_grade(score):
    """Assign valuation grade based on score."""

    if score >= 85:
        return "A+"
    elif score >= 75:
        return "A"
    elif score >= 60:
        return "B"
    elif score >= 45:
        return "C"
    else:
        return "D"

def export_to_excel(df):
    """Export valuation summary to Excel."""

    OUTPUT_DIR.mkdir(exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation Summary"

    headers = [
        "Company",
        "Year",
        "ROE %",
        "Net Margin %",
        "Operating Margin %",
        "Debt/Equity",
        "Interest Coverage",
        "Asset Turnover",
        "FCF",
        "EPS",
        "Book Value",
        "Valuation Score",
        "Grade",
    ]

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid",
    )

    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in df.itertuples(index=False):
        ws.append([
            row.company_name.strip(),
            row.year,
            row.return_on_equity_pct,
            row.net_profit_margin_pct,
            row.operating_profit_margin_pct,
            row.debt_to_equity,
            row.interest_coverage,
            row.asset_turnover,
            row.free_cash_flow_cr,
            row.earnings_per_share,
            row.book_value_per_share,
            row.valuation_score,
            row.grade,
        ])

    for column in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(length + 3, 30)

    ws.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)

    print(f"\nValuation report saved to: {OUTPUT_FILE}")

if __name__ == "__main__":

    df = get_latest_financial_ratios()

    df = calculate_valuation_score(df)

    df["grade"] = df["valuation_score"].apply(assign_grade)

    export_to_excel(df)

    print(df[["company_name", "valuation_score", "grade"]].head())

    print(f"\nCompanies : {len(df)}")