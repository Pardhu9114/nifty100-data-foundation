from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from src.screener.engine import run_preset, load_config

OUTPUT = Path("output/screener_output.xlsx")

green = PatternFill(fill_type="solid", start_color="C6EFCE")
red = PatternFill(fill_type="solid", start_color="FFC7CE")


def export_screeners():

    config = load_config()

    wb = Workbook()

    wb.remove(wb.active)

    for preset in config:

        df = run_preset(preset)

        ws = wb.create_sheet(title=preset[:31])

        ws.append(df.columns.tolist())

        for row in df.itertuples(index=False):
            ws.append(list(row))

        # simple formatting
        for row in ws.iter_rows(min_row=2):

            for cell in row:

                if isinstance(cell.value, (int, float)):

                    if cell.value > 0:
                        cell.fill = green
                    else:
                        cell.fill = red

    OUTPUT.parent.mkdir(exist_ok=True)

    wb.save(OUTPUT)

    print(f"Excel exported -> {OUTPUT}")


if __name__ == "__main__":
    export_screeners()